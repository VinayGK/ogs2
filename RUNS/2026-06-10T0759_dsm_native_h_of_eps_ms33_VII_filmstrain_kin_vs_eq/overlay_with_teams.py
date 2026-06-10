#!/usr/bin/env python3
"""MS33 Model VII inter-team overlay — film_strain_coupling variants.

EXACTLY mirrors the canonical fig_modelVII of
MS33_SUBMISSION_maxwell_2026-06-10/scripts/plot_ms33_interteam_III_IV_VII.py
(teams read from their xlsx submissions: axial stress col 30 vs void ratio
col 58, row-paired; linear axes; domain-averaged ours), with the single BGR
curve replaced by the three film_strain_coupling variants run on the same
free-swell + load/unload protocol (branch dsm_native_h_of_eps @ 7ff8861847).
"""
import glob
import os
import re
import sys

import numpy as np
import openpyxl
import vtk
from vtk.util.numpy_support import vtk_to_numpy
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, "/Users/vinaykumar/tex/cc2024/VK_SB_EURAD_DSM")
try:
    from ms33_provenance_footer import add_provenance_footer
except Exception:
    add_provenance_footer = None

BASE = "/Users/vinaykumar/tex/eurad2_MS34/MSXX/g_Support_Section_Data Collection"
COMMIT = "7ff8861847 (dsm_native_h_of_eps)"
TEAM_COL = {"CTU-CU": "#ff7f0e", "BGE-CU-TUBAF-UFZ": "#17becf"}
VARIANTS = [("off", "#000000", "-"),
            ("kinematic", "#1f77b4", "-"),
            ("equilibrium", "#cc0000", "--")]


def load(p):
    r = vtk.vtkXMLUnstructuredGridReader()
    r.SetFileName(p)
    r.Update()
    return r.GetOutput()


def vii_path(globpat):
    """Domain-averaged (e, axial stress [MPa, compr.+], t[d]) — verbatim from
    the canonical script's vii_path."""
    fs = sorted(glob.glob(globpat),
                key=lambda f: int(re.search(r"_ts_(\d+)_", f).group(1)))
    e, axial, t = [], [], []
    for f in fs:
        g = load(f)
        sig = vtk_to_numpy(g.GetPointData().GetArray("sigma"))
        poro = vtk_to_numpy(g.GetPointData().GetArray("porosity"))
        phi = float(np.mean(poro))
        e.append(phi / (1.0 - phi))
        axial.append(float(np.mean(-sig[:, 1])) / 1e6)
        t.append(float(re.search(r"_t_([0-9.]+)\.vtu", f).group(1)) / 86400.0)
    return np.array(e), np.array(axial), np.array(t)


def team_vii_path(team_file):
    """Verbatim from the canonical script: axial col 30 vs void col 58."""
    wb = openpyxl.load_workbook(team_file, read_only=True, data_only=True)
    ws = wb["Model_VII"]
    axial, e = [], []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        a, v = row[29], row[57]
        if isinstance(a, (int, float)) and isinstance(v, (int, float)):
            axial.append(a)
            e.append(v)
    wb.close()
    return np.array(axial), np.array(e)


def main():
    bge_ax, bge_e = team_vii_path(os.path.join(
        BASE, "BGE-CU-TUBAF-UFZ_DATA/Model_VII/Model_VII_BGE-CU-TUBAF-UFZ.xlsx"))
    ctu_ax, ctu_e = team_vii_path(os.path.join(
        BASE, "CTU-CU_DATA/Model_VII/Model_VII_CTU_CU.xlsx"))

    fig, ax = plt.subplots(figsize=(9.0, 6.4))
    ax.plot(bge_ax, bge_e, "-o", color=TEAM_COL["BGE-CU-TUBAF-UFZ"], ms=3.2,
            lw=1.4, alpha=0.85, label="BGE-CU-TUBAF-UFZ", zorder=3)
    ax.plot(ctu_ax, ctu_e, "-s", color=TEAM_COL["CTU-CU"], ms=3.2, lw=1.4,
            alpha=0.85, label="CTU-CU", zorder=3)

    finals = {}
    for v, col, ls in VARIANTS:
        e, axial, t = vii_path(
            os.path.join(HERE, f"out_{v}",
                         "ms33_modelVII_freeswelling_ts_*_t_*.vtu"))
        ax.plot(axial, e, ls, color=col, lw=2.6,
                label=f"BGR DSM — film_strain={v}", zorder=5)
        ax.plot(axial[-1], e[-1], "*", color=col, ms=15, mec="k", mew=0.7,
                zorder=6)
        finals[v] = (axial[-1], e[-1])

    ax.annotate("team $e_{end}\\approx1.09$", xy=(ctu_ax[-1], ctu_e[-1]),
                xytext=(2.0, 0.92), fontsize=9, color="#555", va="center",
                arrowprops=dict(arrowstyle="->", color="#888", lw=0.8))
    ax.annotate("off $\\approx$ equilibrium\n$e_{end}\\approx%.2f$"
                % finals["off"][1],
                xy=finals["off"], xytext=(1.6, 1.50), fontsize=9,
                color="#cc0000", va="center",
                arrowprops=dict(arrowstyle="->", color="#cc0000", lw=0.8))
    ax.annotate("kinematic $e_{end}\\approx%.2f$" % finals["kinematic"][1],
                xy=finals["kinematic"], xytext=(1.6, 1.28), fontsize=9,
                color="#1f77b4", va="center",
                arrowprops=dict(arrowstyle="->", color="#1f77b4", lw=0.8))

    ax.set_xlabel("axial stress  $\\sigma_{ax}$  (MPa)", fontsize=12)
    ax.set_ylabel("void ratio  $e=\\varphi/(1-\\varphi)$", fontsize=12)
    ax.set_ylim(0.65, 1.58)
    ax.set_title("MS33 Model VII (free swelling) - void ratio vs axial "
                 "stress, load/unload path\n"
                 "film_strain_coupling variants vs BGE & CTU", fontsize=12.5,
                 fontweight="bold")
    ax.grid(True, color="#E0E0E0", lw=0.8, zorder=0)
    ax.legend(fontsize=9.5, loc="upper center", framealpha=1.0,
              edgecolor="#BBB")
    caveat = ("off $\\approx$ equilibrium: the force-balance branch point "
              "never opens at $p\\leq5$ MPa ($\\Pi\\approx40$ MPa) and $n_l$ "
              "is ceiling-pinned.\nkinematic: strained-film geometric "
              "feedback closes part of the over-swelling gap; $K$ NOT "
              "recalibrated (frozen-geometry calibration).")
    ax.text(0.985, 0.025, caveat, transform=ax.transAxes, fontsize=8.0,
            color="#5a2d00", va="bottom", ha="right",
            bbox=dict(boxstyle="round,pad=0.5", fc="#FFEFE0", ec="#cc5500",
                      lw=1.0))
    fig.tight_layout()
    if add_provenance_footer is not None:
        add_provenance_footer(fig, run=HERE, commit=COMMIT)
    out = os.path.join(HERE, "ms33_modelVII_interteam_filmstrain_overlay.png")
    fig.savefig(out, dpi=220)
    print("wrote", out)
    for v, (a, e) in finals.items():
        print(f"  {v:12s} e_end={e:.3f} at sigma_ax={a:.2f} MPa")


if __name__ == "__main__":
    main()
