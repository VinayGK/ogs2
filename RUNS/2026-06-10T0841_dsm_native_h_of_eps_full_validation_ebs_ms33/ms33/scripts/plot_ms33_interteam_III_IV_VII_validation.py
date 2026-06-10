#!/usr/bin/env python3
"""ANCHORS MS33 inter-team comparison overlays for the newly-converged BGR DSM
runs of Models III (gap2mm), IV (pellets) and VII (free-swelling), against the
other EURAD-2 teams.

OUR data (read-only, converged production VTUs; no re-invented numbers):
  III : .../ANCHORS_MS33_ModelIII/conv3_out/final/conv3_diff_final_ts_1443_t_17280000.vtu (+ series)
  IV  : .../runs/conv4_scratch/out_C/ms33_modelIV_pellets_ts_1186_t_17280000.vtu (+ series)
  VII : /tmp/vii_full/ms33_modelVII_freeswelling_ts_*_t_*.vtu  (18 frames, load/unload)

Geometry: 2D axisymmetric, x = radial r [m], y = axial height z [m], z-coord = 0.
Probe map: Top (r=0,z=70mm), Central (r=0,z=40mm), Bottom (r=0,z=10mm).
Mean stress (compression-positive) p = -(sxx+syy+szz)/3.  sigma = [xx,yy,zz,xy].
  radial  ~ sigma_xx (rr),  axial ~ sigma_yy (zz/height).
Void ratio e = phi/(1-phi), phi = porosity (domain-average for the free-swell column).

TEAM data (read-only): Model_{III,IV,V}.xlsx main sheet "Final mean stress" block
(col G, rows 10-14: Top/Central/Bottom), verbatim. Model VII: main-sheet columns
axial-stress (col 29, Top) + void-ratio (col 57) time series, paired by row.

Outputs (220 dpi PNG) to both the tex dir and /tmp.
"""
import os, glob, re, sys
import numpy as np
import vtk
from vtk.util.numpy_support import vtk_to_numpy
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

# Reusable provenance footer (lives alongside this script).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ms33_provenance_footer import add_provenance_footer

OUT_DIRS = ["/Users/vinaykumar/git/ogs/RUNS/_INPROGRESS_full_validation/ms33/figures"]
BASE = "/Users/vinaykumar/tex/eurad2_MS34/MSXX/g_Support_Section_Data Collection"
DPI = 220

# VALIDATION BUNDLE 2026-06-10: adapted copy of the canonical
# MS33_SUBMISSION_maxwell_2026-06-10 overlay generator. ONLY the "ours" inputs,
# output dir, run labels/notes and footer provenance changed; team-reading code
# untouched. Runs: dsm_native_h_of_eps @7ff8861847 (film_strain_coupling OFF =
# bit-for-bit maxwell_conjugate baseline d98f5f8324), binary h_of_eps_20260609.
EXISTING_COMMIT = "7ff8861847"
RUNBASE = "/Users/vinaykumar/git/ogs/RUNS/_INPROGRESS_full_validation/ms33"
import ms33_provenance_footer as _mpf
_mpf._BINARY = "h_of_eps_20260609"

# Distinct team colours (kept stable across panels)
TEAM_COL = {
    "CIMNE-UPC": "#1f77b4", "CTU-CU": "#ff7f0e", "ICL": "#2ca02c",
    "UCLM": "#d62728", "ULIEGE": "#9467bd", "Amphos21": "#8c564b",
    "LEI": "#e377c2", "UBERN": "#7f7f7f", "BGE-CU-TUBAF-UFZ": "#17becf",
}
OURS = "#000000"   # thick black = BGR DSM (ours)
RED  = "#cc0000"   # red diamond markers for ours
LOCS = ["Top", "Central", "Bottom"]

# ---------------------------------------------------------------- VTU helpers
def load(p):
    r = vtk.vtkXMLUnstructuredGridReader(); r.SetFileName(p); r.Update(); return r.GetOutput()

def probe_idx(g, r_mm, z_mm):
    pts = vtk_to_numpy(g.GetPoints().GetData())
    tgt = np.array([r_mm / 1000., z_mm / 1000., 0.0])
    return int(np.argmin(np.linalg.norm(pts - tgt, axis=1)))

def our_final_meanstress(vtu):
    """Return dict Top/Central/Bottom -> mean stress [MPa, compression+]."""
    g = load(vtu); sig = vtk_to_numpy(g.GetPointData().GetArray("sigma"))
    out = {}
    for name, z in zip(LOCS, (70, 40, 10)):
        i = probe_idx(g, 0, z); s = sig[i]
        out[name] = -(s[0] + s[1] + s[2]) / 3.0 / 1e6
    return out

def our_domain_meanstress(vtu):
    """Domain-averaged mean stress [MPa, compression+] over all nodes."""
    g = load(vtu); sig = vtk_to_numpy(g.GetPointData().GetArray("sigma"))
    return float(np.mean(-(sig[:, 0] + sig[:, 1] + sig[:, 2]) / 3.0)) / 1e6

def vii_path(globpat):
    """Domain-averaged (void ratio e, axial stress, mean stress, t[d]) along series."""
    fs = sorted(glob.glob(globpat), key=lambda f: int(re.search(r"_ts_(\d+)_", f).group(1)))
    e, axial, mean, t = [], [], [], []
    for f in fs:
        g = load(f)
        sig = vtk_to_numpy(g.GetPointData().GetArray("sigma"))
        poro = vtk_to_numpy(g.GetPointData().GetArray("porosity"))
        phi = float(np.mean(poro))
        e.append(phi / (1.0 - phi))
        axial.append(float(np.mean(-sig[:, 1])) / 1e6)
        mean.append(float(np.mean(-(sig[:, 0] + sig[:, 1] + sig[:, 2]) / 3.0)) / 1e6)
        t.append(float(re.search(r"_t_([0-9.]+)\.vtu", f).group(1)) / 86400.0)
    return (np.array(e), np.array(axial), np.array(mean), np.array(t))

# ---------------------------------------------------------------- XLSX helpers
def team_final_meanstress(model, want_radial_axial=False):
    """Read 'Final mean stress' (and optionally radial/axial) Top/Central/Bottom
    from each team's Model_{model} main sheet. Skips empty _teamname templates."""
    res = {}
    for d in sorted(glob.glob(os.path.join(BASE, "*_DATA", f"Model_{model}"))):
        for f in glob.glob(os.path.join(d, "*.xlsx")):
            # Skip the shared blank "..._teamname.xlsx" template copies: these carry
            # the same placeholder example numbers in every team dir (III placeholder
            # 5.124/3.297/4.686, IV placeholder 4.944/2.062/3.09), not real
            # submissions. Real submissions are the team-named files.
            if os.path.basename(f).lower().endswith("_teamname.xlsx"):
                continue
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            sn = f"Model_{model}" if f"Model_{model}" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sn]
            if not hasattr(ws, "iter_rows"):
                wb.close(); continue
            block = {}
            for row in ws.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True):
                for j, c in enumerate(row):
                    if not isinstance(c, str):
                        continue
                    lab = c.strip().lower()
                    if lab == "final mean stress":
                        v = [row[j + 1], row[j + 2], row[j + 3]]
                        if all(isinstance(x, (int, float)) for x in v):
                            block["mean"] = v
                    elif want_radial_axial and lab == "final radial stress":
                        v = [row[j + 1], row[j + 2], row[j + 3]]
                        if all(isinstance(x, (int, float)) for x in v):
                            block["radial"] = v
                    elif want_radial_axial and lab == "final axial stress":
                        v = [row[j + 1], row[j + 2], row[j + 3]]
                        if all(isinstance(x, (int, float)) for x in v):
                            block["axial"] = v
            wb.close()
            if "mean" not in block:
                continue
            # team name from the data-folder; normalise label casing
            team = os.path.basename(d.replace(f"/Model_{model}", "")).replace("_DATA", "")
            team = {"AMPOS21": "Amphos21"}.get(team, team)
            res[team] = block
    return res

def team_vii_path(team_file):
    wb = openpyxl.load_workbook(team_file, read_only=True, data_only=True)
    ws = wb["Model_VII"]
    axial, e = [], []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        a, v = row[29], row[57]   # axial stress Top (col 30), void ratio (col 58)
        if isinstance(a, (int, float)) and isinstance(v, (int, float)):
            axial.append(a); e.append(v)
    wb.close()
    return np.array(axial), np.array(e)

def savefig(fig, name, run=None, commit=EXISTING_COMMIT):
    # Stamp a provenance footer (called after the per-figure tight_layout, so the
    # reserved bottom margin survives). commit defaults to the pre-Option-B build
    # the existing figures came from; pass commit=None to auto-stamp current HEAD.
    if run is not None:
        add_provenance_footer(fig, run=run, commit=commit,
                              branch="dsm_native_h_of_eps", date="2026-06-10")
    paths = []
    for d in OUT_DIRS:
        os.makedirs(d, exist_ok=True)
        p = os.path.join(d, name)
        fig.savefig(p, dpi=DPI, bbox_inches="tight")
        paths.append(p)
    plt.close(fig)
    return paths

# ================================================================ FIGURE 1: Model III
def fig_modelIII():
    teams = team_final_meanstress("III")
    ours = our_final_meanstress(
        RUNBASE + "/out_modelIII_gap2mm/"
        "ms33_modelIII_gap2mm_ts_438_t_17280000.000000.vtu")
    BGR_REF = 14.16   # no-gap BGR reference (saturated, Dixon dd1600)
    x = np.arange(len(LOCS))
    fig, ax = plt.subplots(figsize=(7.2, 4.6))
    for team, blk in sorted(teams.items()):
        ax.plot(x, blk["mean"], "o", ms=8, color=TEAM_COL.get(team, "#555"),
                label=team, alpha=0.9, zorder=3)
    ax.plot(x, [ours[l] for l in LOCS], "-D", color=OURS, mfc=RED, mec=OURS,
            mew=1.0, ms=11, lw=2.6, label="BGR DSM (ours)", zorder=5)
    ax.axhline(BGR_REF, color="#444", ls="--", lw=1.3, zorder=2)
    ax.text(-0.35, BGR_REF + 0.25, "BGR no-gap reference (14.16 MPa)", va="bottom",
            ha="left", fontsize=8, color="#444")
    ax.set_xticks(x); ax.set_xticklabels(LOCS, fontsize=11)
    ax.set_xlim(-0.5, 2.5)
    ax.set_ylim(1.0, 19.5)
    ax.set_ylabel("final mean (effective) stress  $p$  (MPa)", fontsize=11)
    ax.set_title("MS33 Model III (2 mm gap) - final mean stress at Top / Central / Bottom\n"
                 "corrected BGR DSM vs EURAD-2 teams", fontsize=11.5, fontweight="bold")
    ax.grid(True, axis="y", color="#E0E0E0", lw=0.8, zorder=0)
    ax.legend(fontsize=8.5, loc="upper center", ncol=3, framealpha=1.0, edgecolor="#BBB")
    note = ("BGR DSM: validation re-run 2026-06-10, dsm_native_h_of_eps @7ff8861847\n"
            "(coupling OFF = maxwell_conjugate baseline), completed t=200 d.")
    ax.text(0.015, 0.025, note, transform=ax.transAxes, fontsize=7.6, color="#333",
            va="bottom", ha="left",
            bbox=dict(boxstyle="round,pad=0.4", fc="#FFF8E1", ec="#D4A017", lw=0.8))
    fig.tight_layout()
    return savefig(fig, "ms33_modelIII_interteam.png",
                   run="RUNS/_INPROGRESS_full_validation/ms33/out_modelIII_gap2mm (ts_438, t=200 d)"), ours, teams

# ================================================================ FIGURE 2: Model IV
def fig_modelIV():
    teams = team_final_meanstress("IV")
    # IV data source — VALIDATION RE-RUN 2026-06-10 (h_of_eps binary, coupling
    # OFF): ms33_modelIV_pellets.prj completed the full 200 d (t=17280000 s,
    # ts_636, EXIT=0). Latest frame == final frame; no partial-run caveat.
    iv_latest = sorted(
        glob.glob(RUNBASE + "/out_modelIV_pellets/ms33_modelIV_pellets_ts_*_t_*.vtu"),
        key=lambda f: float(re.search(r"_t_([0-9.]+)\.vtu", f).group(1)))
    iv_src = iv_latest[-1]   # final 200-d frame (ts_636_t_17280000)
    iv_run = ("RUNS/_INPROGRESS_full_validation/ms33/out_modelIV_pellets "
              "(ts_636, t=200 d, completed)")
    ours = our_final_meanstress(iv_src)
    domain = our_domain_meanstress(iv_src)
    x = np.arange(len(LOCS))
    fig, ax = plt.subplots(figsize=(7.2, 4.6))
    for team, blk in sorted(teams.items()):
        ax.plot(x, blk["mean"], "o", ms=8, color=TEAM_COL.get(team, "#555"),
                label=team, alpha=0.9, zorder=3)
    ax.plot(x, [ours[l] for l in LOCS], "-D", color=OURS, mfc=RED, mec=OURS,
            mew=1.0, ms=11, lw=2.6, label="BGR DSM (ours)", zorder=5)
    ax.axhline(domain, color="#444", ls=":", lw=1.1, zorder=2)
    ax.text(-0.35, domain + 0.2, "ours: domain mean %.2f MPa" % domain, va="bottom",
            ha="left", fontsize=8, color="#444")
    ax.set_xticks(x); ax.set_xticklabels(LOCS, fontsize=11)
    ax.set_xlim(-0.5, 2.5)
    ax.set_ylim(0.0, 13.5)
    ax.set_ylabel("final mean (effective) stress  $p$  (MPa)", fontsize=11)
    ax.set_title("MS33 Model IV (clay-pellet mixture) - final mean stress at Top / Central / Bottom\n"
                 "corrected BGR DSM vs EURAD-2 teams", fontsize=11.5, fontweight="bold")
    ax.grid(True, axis="y", color="#E0E0E0", lw=0.8, zorder=0)
    ax.legend(fontsize=8.5, loc="lower center", ncol=4, framealpha=1.0, edgecolor="#BBB")
    note = ("Model IV: validation re-run 2026-06-10, completed t=200 d\n"
            "(h_of_eps @7ff8861847, coupling OFF).\n"
            "DSM Top/Central/Bottom = %.2f / %.2f / %.2f MPa (domain %.1f MPa)."
            % (ours["Top"], ours["Central"], ours["Bottom"], domain))
    ax.text(0.985, 0.965, note, transform=ax.transAxes, fontsize=7.6, color="#333",
            va="top", ha="right",
            bbox=dict(boxstyle="round,pad=0.4", fc="#E8F4FF", ec="#3a7bbf", lw=0.8))
    fig.tight_layout()
    return savefig(fig, "ms33_modelIV_interteam.png", run=iv_run), ours, teams

# ================================================================ FIGURE 3: Model VII
def fig_modelVII():
    e, axial, mean, t = vii_path(
        RUNBASE + "/out_modelVII_freeswelling/ms33_modelVII_freeswelling_ts_*_t_*.vtu")
    bge_ax, bge_e = team_vii_path(os.path.join(BASE,
        "BGE-CU-TUBAF-UFZ_DATA/Model_VII/Model_VII_BGE-CU-TUBAF-UFZ.xlsx"))
    ctu_ax, ctu_e = team_vii_path(os.path.join(BASE,
        "CTU-CU_DATA/Model_VII/Model_VII_CTU_CU.xlsx"))
    fig, ax = plt.subplots(figsize=(9.0, 6.4))
    # teams: void ratio (y) vs axial stress (x)
    ax.plot(bge_ax, bge_e, "-o", color=TEAM_COL["BGE-CU-TUBAF-UFZ"], ms=3.2, lw=1.4,
            alpha=0.85, label="BGE-CU-TUBAF-UFZ", zorder=3)
    ax.plot(ctu_ax, ctu_e, "-s", color=TEAM_COL["CTU-CU"], ms=3.2, lw=1.4,
            alpha=0.85, label="CTU-CU", zorder=3)
    # ours: load/unload path, ordered by time (already sorted)
    ax.plot(axial, e, "-D", color=OURS, mfc=RED, mec=OURS, mew=0.9, ms=7.5, lw=2.6,
            label="BGR DSM (ours, load/unload path)", zorder=5)
    # mark unload start (max axial) and final
    imax = int(np.argmax(axial))
    ax.annotate("our load peak\n($\\sigma_{ax}=%.1f$ MPa)" % axial[imax],
                xy=(axial[imax], e[imax]),
                xytext=(axial[imax] - 1.6, e[imax] + 0.07), fontsize=8.5, color="#333",
                ha="center",
                arrowprops=dict(arrowstyle="->", color="#777", lw=0.8))
    ax.plot(axial[-1], e[-1], "*", color=RED, ms=17, mec=OURS, mew=0.8, zorder=6)
    ax.annotate("ours $e_{end}\\approx%.2f$ (over-swelled)" % e[-1],
                xy=(axial[-1], e[-1]), xytext=(1.2, e[-1] + 0.02),
                fontsize=9, color=RED, va="center",
                arrowprops=dict(arrowstyle="->", color=RED, lw=0.8))
    ax.annotate("team $e_{end}\\approx1.09$", xy=(ctu_ax[-1], ctu_e[-1]),
                xytext=(2.0, 0.92), fontsize=9, color="#555", va="center",
                arrowprops=dict(arrowstyle="->", color="#888", lw=0.8))
    ax.set_xlabel("axial stress  $\\sigma_{ax}$  (MPa)", fontsize=12)
    ax.set_ylabel("void ratio  $e=\\varphi/(1-\\varphi)$", fontsize=12)
    ax.set_ylim(0.65, 1.55)
    ax.set_title("MS33 Model VII (free swelling) - void ratio vs axial stress, load/unload path\n"
                 "corrected BGR DSM vs BGE & CTU", fontsize=12.5, fontweight="bold")
    ax.grid(True, color="#E0E0E0", lw=0.8, zorder=0)
    ax.legend(fontsize=9.5, loc="upper center", framealpha=1.0, edgecolor="#BBB")
    caveat = ("Validation re-run 2026-06-10 (h_of_eps @7ff8861847, coupling OFF):\n"
              "our $e_{end}=%.2f$; team end-state $e\\approx1.09$ (numbers side by side,\n"
              "no agreement claim implied)." % e[-1])
    ax.text(0.985, 0.025, caveat, transform=ax.transAxes, fontsize=8.4, color="#5a2d00",
            va="bottom", ha="right",
            bbox=dict(boxstyle="round,pad=0.5", fc="#FFEFE0", ec="#cc5500", lw=1.0))
    fig.tight_layout()
    return savefig(fig, "ms33_modelVII_interteam.png",
                   run="RUNS/_INPROGRESS_full_validation/ms33/out_modelVII_freeswelling (full load/unload)"), (axial, e), {"BGE": (bge_ax, bge_e), "CTU": (ctu_ax, ctu_e)}

# ================================================================ FIGURE 4: summary
def fig_summary(ours_III, ours_IV):
    """Final mean stress per model (III, IV, V) - ours vs team spread.
    Model V is team-only context (no converged BGR V here); ours shown for III & IV."""
    teamsV = team_final_meanstress("V")
    teamsIII = team_final_meanstress("III")
    teamsIV = team_final_meanstress("IV")
    models = ["III", "IV", "V"]
    team_sets = {"III": teamsIII, "IV": teamsIV, "V": teamsV}
    our_central = {"III": ours_III["Central"], "IV": ours_IV["Central"]}
    fig, ax = plt.subplots(figsize=(7.2, 4.8))
    for k, m in enumerate(models):
        # all team Central values
        vals = [blk["mean"][1] for blk in team_sets[m].values()]
        names = list(team_sets[m].keys())
        jitter = np.linspace(-0.12, 0.12, len(vals)) if len(vals) > 1 else [0]
        for v, nm, jx in zip(vals, names, jitter):
            ax.plot(k + jx, v, "o", ms=8, color=TEAM_COL.get(nm, "#777"), alpha=0.9, zorder=3)
        # team spread bar (min-max) + mean tick
        ax.plot([k, k], [min(vals), max(vals)], color="#999", lw=1.0, zorder=1)
        ax.plot(k, np.mean(vals), "_", ms=22, color="#555", mew=2.0, zorder=2)
        if m in our_central:
            ax.plot(k, our_central[m], "D", color=RED, mec=OURS, mew=1.2, ms=13, zorder=6)
            # Flag Model IV's diamond as the per-material 100-d PARTIAL (lower bound):
            # the value is still rising, so it is not directly comparable to the
            # equilibrium team / Model III points.
            if m == "IV":
                ax.annotate("per-material K\n100-d partial\n(lower bound)",
                            xy=(k, our_central[m]),
                            xytext=(k + 0.18, our_central[m] + 1.6),
                            fontsize=7.6, color=RED, va="bottom", ha="left",
                            arrowprops=dict(arrowstyle="->", color=RED, lw=0.8))
    # legend
    handles = [Line2D([], [], marker="o", ls="none", color="#777", label="team (Central)"),
               Line2D([], [], marker="_", ls="none", color="#555", mew=2, label="team mean"),
               Line2D([], [], marker="D", ls="none", color=RED, mec=OURS, mew=1.2,
                      ms=11, label="BGR DSM (ours)")]
    ax.legend(handles=handles, fontsize=9, loc="upper left", framealpha=1.0, edgecolor="#BBB")
    ax.set_xticks(range(len(models)))
    ax.set_xticklabels([f"Model {m}" for m in models], fontsize=11)
    ax.set_xlim(-0.5, len(models) - 0.5)
    ax.set_ylabel("final Central mean stress  $p$  (MPa)", fontsize=11)
    ax.set_title("MS33 final Central mean stress per model - BGR DSM vs EURAD-2 team spread",
                 fontsize=11.5, fontweight="bold")
    ax.grid(True, axis="y", color="#E0E0E0", lw=0.8, zorder=0)
    ax.text(0.985, 0.025,
            "Model IV: per-material K, 100-d partial (stalled 137 d; lower bound).\n"
            "Model V: team-only context (no converged BGR V in this set).",
            transform=ax.transAxes, fontsize=7.6, color="#555", va="bottom", ha="right",
            bbox=dict(boxstyle="round,pad=0.35", fc="#F2F2F2", ec="#AAA", lw=0.7))
    fig.tight_layout()
    return savefig(fig, "ms33_all_models_final_meanstress.png",
                   run="III conv3_out/final; IV iv_permat per-material K 100-d PARTIAL "
                       "(stalled 137d); V team-only"), team_sets

# ================================================================ main
if __name__ == "__main__":
    p1, o3, t3 = fig_modelIII()
    p2, o4, t4 = fig_modelIV()
    p3, ours7, teams7 = fig_modelVII()
    # fig_summary intentionally NOT regenerated for the validation bundle:
    # only the three per-model inter-team overlays are requested
    # (confirmed overlay spec 2026-06-10; summary bar chart excluded).

    print("\n===== FIGURES WRITTEN =====")
    for p in (p1 + p2 + p3):
        print("  ", p)

    print("\n===== OUR VALUES (Central) =====")
    print("  III  Top/Central/Bottom = %.2f / %.2f / %.2f MPa" %
          (o3["Top"], o3["Central"], o3["Bottom"]))
    print("  IV   Top/Central/Bottom = %.2f / %.2f / %.2f MPa" %
          (o4["Top"], o4["Central"], o4["Bottom"]))
    ax7, e7 = ours7
    print("  VII  e_end=%.3f, e_max=%.3f, load-peak axial=%.2f MPa" %
          (e7[-1], e7.max(), ax7[ax7.argmax()] if len(ax7) else float("nan")))

    print("\n===== TEAM Central final-mean-stress (MPa) =====")
    for mdl, ts in (("III", t3), ("IV", t4)):
        print(f"  Model {mdl}:")
        for team, blk in sorted(ts.items()):
            print(f"     {team:20s} Top/Cen/Bot = %.2f / %.2f / %.2f" %
                  (blk['mean'][0], blk['mean'][1], blk['mean'][2]))
    print("\n  VII teams e_end: BGE=%.3f  CTU=%.3f" %
          (teams7["BGE"][1][-1], teams7["CTU"][1][-1]))
